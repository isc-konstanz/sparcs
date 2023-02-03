# -*- coding: utf-8 -*-
"""
    th-e-yield.system
    ~~~~~~~~~~~~~~~~~
    
    
"""
import os
import json
import logging
import pandas as pd
import datetime as dt
import th_e_core
from th_e_core import Component
from th_e_core.configs import Configurations, ConfigurationUnavailableException
from th_e_core.weather import Weather
from typing import Dict
from .pv import PVSystem
from .model import Model
from .location import Location

logger = logging.getLogger(__name__)

AC_P = 'Power [W]'
AC_E = 'Energy yield [kWh]'
AC_Y = 'Specific yield [kWh/kWp]'
DC_P = 'Power (DC) [W]'
DC_E = 'Energy yield (DC) [kWh]'


class System(th_e_core.System):

    def __configure__(self, configs):
        super().__configure__(configs)
        data_dir = configs.dirs.data
        if not os.path.exists(data_dir):
            os.makedirs(data_dir, exist_ok=True)

        self._results_json = os.path.join(data_dir, 'results.json')
        self._results_excel = os.path.join(data_dir, 'results.xlsx')
        self._results_csv = os.path.join(data_dir, 'results.csv')
        self._results_dir = os.path.join(data_dir, 'results')
        if not os.path.exists(self._results_dir):
            os.makedirs(self._results_dir)

    def __activate__(self, components: Dict[str, Component], configs: Configurations) -> None:
        super().__activate__(components, configs)
        self.weather = Weather.read(self)

    def __init_location__(self, configs: Configurations) -> Location:
        # FIXME: location necessary for for weather instantiation, but called afterwards here
        # if isinstance(self.weather, TMYWeather):
        #     return Location.from_tmy(self.weather.meta)
        # elif isinstance(self.weather, EPWWeather):
        #     return Location.from_epw(self.weather.meta)

        return Location(configs.getfloat('Location', 'latitude'),
                        configs.getfloat('Location', 'longitude'),
                        timezone=configs.get('Location', 'timezone', fallback='UTC'),
                        altitude=configs.getfloat('Location', 'altitude', fallback=None),
                        country=configs.get('Location', 'country', fallback=None),
                        state=configs.get('Location', 'state', fallback=None))

    def __cmpt_types__(self):
        return super().__cmpt_types__('solar', 'array')

    # noinspection PyShadowingBuiltins
    def __cmpt__(self, configs: Configurations, type: str) -> Component:
        if type in ['pv', 'solar', 'array']:
            return PVSystem(self, configs)

        return super().__cmpt__(configs, type)

    def __call__(self) -> pd.DataFrame:
        logger.info("Starting simulation for system: %s", self.name)
        start = dt.datetime.now(self.location.timezone)

        weather = self.weather.get(start=start)
        if 'precipitable_water' not in weather.columns or weather['precipitable_water'].sum() == 0:
            from pvlib.atmosphere import gueymard94_pw
            weather['precipitable_water'] = gueymard94_pw(weather['temp_air'], weather['relative_humidity'])
        if 'albedo' in weather.columns and weather['albedo'].sum() == 0:
            weather.drop('albedo', axis=1, inplace=True)

        result = pd.DataFrame(columns=['p_ac', 'p_dc'], index=weather.index).fillna(0)
        result.index.name = 'time'
        results = {}
        try:
            progress = Progress(len(self) + 1, file=self._results_json)

            for key, cmpt in self.items():
                if cmpt.type == 'pv':
                    model = Model.read(cmpt)
                    data = model(weather)

                    if self._database is not None:
                        self._database.persist(data)

                    result[['p_ac', 'p_dc']] += data[['p_ac', 'p_dc']].abs()
                    results[key] = data
                    progress.update()

            results_json = self.evaluate(results, weather)

            with open(self._results_json, 'w', encoding='utf-8') as f:
                json.dump(results_json, f, ensure_ascii=False, indent=4)

        except Exception as e:
            with open(self._results_json, 'w', encoding='utf-8') as f:
                import traceback
                results_json = {
                    'status': 'error',
                    'message': str(e),
                    'error': type(e).__name__,
                    'trace': traceback.format_exc()
                }
                json.dump(results_json, f, ensure_ascii=False, indent=4)

            raise e

        logger.info('Simulation complete')
        logger.debug('Simulation complete in %i seconds',
                     (dt.datetime.now(self.location.timezone) - start).total_seconds())

        return pd.concat([result, weather], axis=1)

    def evaluate(self, results, weather):
        json_results = {}
        json_results.update(self._evaluate_yield(results, weather))

        return json_results

    def _evaluate_yield(self, results, weather):
        hours = pd.Series(weather.index, index=weather.index)
        hours = round((hours - hours.shift(1)).fillna(method='bfill').dt.total_seconds() / 3600)

        results_total_columns = [AC_P, AC_E, AC_Y, DC_P]
        results_total = pd.DataFrame(index=weather.index, columns=[AC_P, DC_P]).fillna(0)
        results_total.index.name = 'Time'
        results_kwp = 0
        for key, system in self.items():
            result = results[key]
            results_kwp += system.power_max / 1000

            results_total[[AC_P, DC_P]] += result[['p_ac', 'p_dc']].values

        results_total[AC_E] = results_total[AC_P] / 1000 * hours
        results_total[AC_Y] = results_total[AC_E] / results_kwp

        results_total[DC_E] = results_total[DC_P] / 1000 * hours

        results_total = results_total.dropna(axis='index', how='all')

        yield_specific = round(results_total[AC_Y].sum(), 2)
        yield_energy = round(results_total[AC_E].sum(), 2)

        ghi = round((weather['ghi'] / 1000 * hours).sum(), 2)
        dhi = round((weather['dhi'] / 1000 * hours).sum(), 2)

        results_total = results_total[results_total_columns]

        results_summary_columns = [AC_E, AC_Y]
        results_summary_data = [yield_energy, yield_specific]

        results_summary_columns.append('GHI [kWh/m^2]')
        results_summary_data.append(ghi)

        results_summary_columns.append('DHI [kWh/m^2]')
        results_summary_data.append(dhi)

        results_summary = pd.DataFrame(data=[results_summary_data], columns=results_summary_columns)

        self._write_csv(results_summary, results_total, results)
        self._write_excel(results_summary, results_total, results)
        return {
                   'status': 'success',
                   'yield_energy': yield_energy,
                   'yield_specific': yield_specific
               }

    def _write_csv(self, results_summary, results_total, results):
        for key, configs in self.items():
            configs_name = configs.name
            for configs_type in self.get_types():
                configs_name = configs_name.replace(configs_type, '')
            if len(configs_name) < 1:
                configs_name += str(list(self.values()).index(configs) + 1)

            results[key].to_csv(os.path.join(self._results_dir, 'results_'+configs_name+'.csv'), encoding='utf-8-sig')

        results_total.to_csv(os.path.join(self._results_dir, 'results.csv'), encoding='utf-8-sig')
        results_summary.to_csv(self._results_csv, encoding='utf-8-sig', index=False)

        return results

    def _write_excel(self, results_summary, results_total, results):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Border, Font, Side

            border_side = Side(border_style=None)
            border = Border(top=border_side,
                            right=border_side,
                            bottom=border_side,
                            left=border_side)

            results_book = Workbook()
            results_writer = pd.ExcelWriter(self._results_excel, engine='openpyxl')
            results_writer.book = results_book
            results_summary.to_excel(results_writer, sheet_name='Summary', float_format="%.2f", encoding='utf-8-sig')
            results_book['Summary'].delete_cols(1, 1)
            results_book.remove_sheet(results_book.active)
            results_book.active = 0

            results_total.tz_localize(None).to_excel(results_writer, sheet_name=self.name, encoding='utf-8-sig')

            for key, configs in self.items():
                configs_name = configs.name
                for configs_type in self.get_types():
                    configs_name = configs_name.replace(configs_type, '')
                if len(configs_name) < 1:
                    configs_name += str(list(self.values()).index(configs) + 1)

                results[key].tz_localize(None).to_excel(results_writer,
                                                        sheet_name=(self.name + ' ' + configs_name).title(),
                                                        encoding='utf-8-sig')

            results_header_font = Font(name="Calibri Light", size=12, color='333333')
            for result_sheet in results_book:
                result_index_width = 0
                for result_row in result_sheet:
                    result_row[0].border = border
                    result_index_width = max(result_index_width, len(str(result_row[0].value)))

                result_sheet.column_dimensions[get_column_letter(1)].width = result_index_width + 2

                for result_column in range(1, len(result_sheet[1])):
                    result_column_width = len(str(result_sheet[1][result_column].value))
                    result_sheet.column_dimensions[get_column_letter(result_column+1)].width = result_column_width + 2
                    result_sheet[1][result_column].font = results_header_font
                    result_sheet[1][result_column].border = border

            results_book.save(self._results_excel)
            results_writer.close()

        except ImportError as e:
            logger.debug("Unable to write excel file for {} of system {}: {}".format(
                os.path.abspath(self._results_excel), self.name, str(e)))


class Progress:

    def __init__(self, total, value=0, file=None):
        self._file = file
        self._total = total
        self._value = value

    def update(self):
        self._value += 1
        self._update(self._value)

    def _update(self, value):
        progress = value / self._total * 100
        if progress % 1 <= 1 / self._total * 100 and self._file is not None:
            with open(self._file, 'w', encoding='utf-8') as f:
                results = {
                    'status': 'running',
                    'progress': int(progress)
                }
                json.dump(results, f, ensure_ascii=False, indent=4)

